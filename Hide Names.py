import openpyxl

# Path to the Excel file
file_path = "C:\\Users\\kerem\\PokemonDatabase\\Code\\Edit Datasheet Exel\\Prompts and Images.xlsx"

# Load the workbook and select Sheet1
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Sheet']

# Loop through the rows
for row in range(3, 13):  # Rows are 1-indexed in openpyxl
    word_to_replace = sheet.cell(row=row, column=2).value
    if word_to_replace and isinstance(word_to_replace, str):
        cell_value_in_col_I = sheet.cell(row=row, column=3).value
        if cell_value_in_col_I and isinstance(cell_value_in_col_I, str):
            # Replace the word in column I with "It"
            sheet.cell(row=row, column=3).value = cell_value_in_col_I.replace(word_to_replace, "It")

# Save the workbook with the changes
workbook.save(file_path)
