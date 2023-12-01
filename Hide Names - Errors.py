import openpyxl

# Path to the Excel file
file_path = "C:\\Users\\kerem\\PokemonDatabase\\Code\\Edit Datasheet Exel\\Prompts and Images.xlsx"
# Load the workbook and select Sheet1
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Sheet']

# Iterate over the rows from 3 to 153
for row_num in range(3, 13):  # Rows are 1-indexed in openpyxl
    word_to_replace = sheet.cell(row=row_num, column=2).value
    
    # If the word exists and is a string, find and replace its occurrences in column I
    if word_to_replace and isinstance(word_to_replace, str):
        for i_row in range(3, 112):
            cell_value_in_col_I = sheet.cell(row=i_row, column=3).value
            if cell_value_in_col_I and isinstance(cell_value_in_col_I, str):
                # Replace the word in column I with "ERRORFIX"
                sheet.cell(row=i_row, column=3).value = cell_value_in_col_I.replace(word_to_replace, "ERRORFIX")

# Save the workbook with the changes in a new file named "output2.xlsx"
output_path = "C:\\Users\\kerem\\PokemonDatabase\\Code\\Edit Datasheet Exel\\Prompts and Images Name Fix.xlsx"
workbook.save(output_path)
