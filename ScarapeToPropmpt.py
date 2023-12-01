import openpyxl
import os

# Load the 'Test Dataset.xlsx' workbook
test_dataset_path = "C:\\Users\\kerem\\PokemonDatabase\\Code\\Text to Prompt Code\\Text Dataset.xlsx"

test_dataset_wb = openpyxl.load_workbook(test_dataset_path)
test_dataset_sheet = test_dataset_wb.active

# Load or create 'Prompts and Images.xlsx' workbook
output_path = "C:\\Users\\kerem\\PokemonDatabase\\Code\\Text to Prompt Code\\Prompts and Images.xlsx"
output_wb = openpyxl.load_workbook(output_path) if os.path.exists(output_path) else openpyxl.Workbook()
output_sheet = output_wb.active

# Loop through rows 3 to 12
for n in range(3, 13):
    Input_Name = test_dataset_sheet.cell(row=n, column=2).value
    Input_Short_Look = test_dataset_sheet.cell(row=n, column=3).value
    Input_Type = test_dataset_sheet.cell(row=n, column=4).value
    Input_Egg_Group = test_dataset_sheet.cell(row=n, column=5).value
    Input_Height = test_dataset_sheet.cell(row=n, column=6).value
    Input_Weight = test_dataset_sheet.cell(row=n, column=7).value
    Input_Long_Description = test_dataset_sheet.cell(row=n, column=8).value

    if Input_Egg_Group != "No Eggs Discovered":
        Text_Prompt = f"Generate a Pokemon that looks like {Input_Egg_Group} {Input_Short_Look} with types of {Input_Type}. It has a height of approximately {Input_Height} and weight of approximately {Input_Weight}. {Input_Long_Description}."
        Image_Location = Input_Name
    else:
        Text_Prompt = f"Generate a Pokemon that looks like {Input_Short_Look} with types of {Input_Type}. It has a height of approximately {Input_Height} and weight of approximately {Input_Weight}. {Input_Long_Description}."
        Image_Location = Input_Name

    # Save to 'Prompts and Images.xlsx'
    output_sheet.cell(row=n, column=2).value = Image_Location  # Column B
    output_sheet.cell(row=n, column=3).value = Text_Prompt     # Column C

# Save changes to the 'Prompts and Images.xlsx'
output_wb.save(output_path)
print("Operation completed successfully.")
